import csv

from openpyxl.reader.excel import load_workbook


def convert_xlsx_to_csv(xlsx_file: str, csv_file: str) -> None:
    wb = load_workbook(xlsx_file)
    sheet = wb.active

    with open(csv_file, mode="w", newline='', encoding="utf-8") as file:
        writer = csv.writer(file, delimiter=';')

        for row in sheet.iter_rows(values_only=True):
            writer.writerow(row)
